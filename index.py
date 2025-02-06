import os
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import random
import tkinter as tk
from tkinter import *
from tkinter import messagebox
from customtkinter import *
from customtkinter import CTkToplevel
import tkinter.scrolledtext as st 
from PIL import Image, ImageTk
import threading

# Fungsi buat GUI  
def create_widgets():
    # Logo
    my_image = Image.open("./assets/SMG-logo.png")
    my_image_resized = my_image.resize((103, 67))
    my_image_tk = ImageTk.PhotoImage(my_image_resized)
    root.image_label = CTkLabel(root, image=my_image_tk, text="",bg_color='transparent')
    root.image_label.place(relx=0.5, rely=0.14, anchor="center")  # Tengah secara horizontal dan sedikit di atas

    add_img = Image.open("./assets/add.png")
    add_img_resized = add_img.resize((22, 22))
    add_img_tk = ImageTk.PhotoImage(add_img_resized)

    # Label-input name
    root.labelname = CTkLabel(root, text="Input name:", bg_color='transparent')
    root.labelname.place(relx=0.09, rely=0.30, anchor="w")

    root.entryname = CTkEntry(root, width=120, textvariable=namevar)
    root.entryname.place(relx=0.09, rely=0.38, anchor="w")
    root.entryname.bind("<Return>", log_data)

    root.addButton = CTkButton(root, text='', image=add_img_tk, command=log_data, fg_color='transparent', hover_color="", width=10)
    root.addButton.place(relx=0.42, rely=0.38, anchor="w")

    # Log Label
    root.logLabel = CTkLabel(root, text="Names:")
    root.logLabel.place(relx=0.09, rely=0.47, anchor="w")

    root.logFrame = st.ScrolledText(root, wrap=tk.WORD,  font=("Montserrat", 16)) 
    root.logFrame.place(relx=0.09, rely=0.68, anchor="w", relwidth=0.45, relheight=0.32,)
    root.logFrame.configure(state ='disabled') 

    # Label-input date
    root.labeldate = CTkLabel(root, text="Number of Days:")
    root.labeldate.place(relx=0.58, rely=0.30, anchor="w")

    root.datecomBox = CTkComboBox(root, values=['28', '29', '30', '31'], width=120)
    root.datecomBox.place(relx=0.58, rely=0.38, anchor="w")

    # Label-input first day of the month
    root.labeldays = CTkLabel(root, text="First Day:")
    root.labeldays.place(relx=0.58, rely=0.47, anchor="w")

    root.daycomBox = CTkComboBox(root, values=['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'], width=120)
    root.daycomBox.place(relx=0.58, rely=0.55, anchor="w")

    # rules Button
    root.rulesBTN = CTkButton(root, text="Settings", command=top_level_win, width=120, fg_color="gray", corner_radius=32, text_color='white')
    root.rulesBTN.place(relx=0.58, rely=0.71, anchor="w")

    # clear Button
    root.clearBTN = CTkButton(root, text="Clear", command=clear_data, width=120, fg_color="red", corner_radius=32, text_color='white')
    root.clearBTN.place(relx=0.58, rely=0.80, anchor="w")

    # generate Button
    root.GenerateBTN = CTkButton(root, text="Generate", command=start_generate_data_thread, width=300, fg_color="green", corner_radius=32, text_color='white')
    root.GenerateBTN.place(relx=0.5, rely=0.92, anchor="center")

# button add name func
import tkinter as tk
from tkinter import messagebox

# Fungsi untuk log data
def log_data(event=None):
    input_name = root.entryname.get().strip()

    if input_name != '':  
        # Memastikan input hanya berisi huruf dan mengubah huruf pertama menjadi kapital
        if input_name.isalpha():
            input_name = input_name.capitalize()
            if input_name in inputted_names:
                messagebox.showerror("ERROR", f"Name '{input_name}' Already inserted!")
            else:
                inputted_names.append(input_name)
                root.logFrame.config(state='normal')
                root.logFrame.insert(tk.END, input_name + '\n') 
                root.logFrame.config(state='disabled')
                root.entryname.delete(0, tk.END)
        else:
            messagebox.showerror("ERROR", "Please input a valid name containing only letters!")
    else:
        messagebox.showerror("ERROR", "Please input a name!")


# Fungsi clear data
def clear_data():
    if hasattr(root, 'new_window') and root.new_window.winfo_exists():
        messagebox.showerror("WARNING", "Save or close settings first!", parent=root.new_window)
    else:
        root.logFrame.config(state='normal')
        root.logFrame.delete('1.0', tk.END)
        root.logFrame.config(state='disabled')
        inputted_names.clear()
        saved_data.clear()

# Fungsi untuk memulai thread untuk generate data
def start_generate_data_thread():
    thread = threading.Thread(target=generate_data)
    thread.start()

def generate_data():
    
    if not saved_data:  # Jika ada data cuti
        withoutrulescuti()
    else:
        with_rulescuti()
    root.attributes("-topmost", True)
    messagebox.showinfo("Success", "Shift Generated!", parent=root)
    root.attributes("-topmost", False)

# Fungsi dengan aturan cuti yang lebih adil
def with_rulescuti():
    global max_days
    if hasattr(root, 'new_window') and root.new_window.winfo_exists():
        messagebox.showerror("WARNING", "Save or close settings first!", parent=root.new_window)
    else:
        max_days = int(root.datecomBox.get())
        start_day = root.daycomBox.get()

        days_of_week = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        start_index = days_of_week.index(start_day)
        week_days = days_of_week[start_index:] + days_of_week[:start_index]

        if inputted_names:
            shift_data = []
            total_names = len(inputted_names)

            # Dictionary untuk menghitung jumlah shift
            shift_count = {name: {"Malam": 0, "Pagi": 0, "Sore": 0} for name in inputted_names}
            weekly_shift_count = {name: {"Malam": [], "Pagi": [], "Sore": []} for name in inputted_names}

            # Hitung total shift yang dibutuhkan per orang
            total_shift_count = {name: 0 for name in inputted_names}
            total_shift_per_person = max_days * 3 // total_names  # Jumlah shift rata-rata per orang

            # Loop untuk setiap hari
            for day in range(max_days):
                current_day = week_days[day % len(week_days)]

                # Ambil list nama yang tersedia (tidak cuti pada hari ini)
                available_names = []
                leave_names = []  # Simpan nama yang cuti
                for name in inputted_names:
                    leave_days = saved_data.get(name, {}).get('days', [])
                    if str(day + 1) not in leave_days:  # Jika tidak ada di daftar cuti
                        available_names.append(name)
                    else:
                        leave_names.append(name)  # Nama orang yang cuti

                # Jika tidak cukup orang yang tersedia, tambahkan orang sebagai backup
                if len(available_names) < 3:
                    additional_names = [name for name in inputted_names if name not in available_names]
                    while len(available_names) < 3:
                        available_names.append(additional_names.pop(0))

                # Pastikan ada cukup orang yang tersedia, jika tidak, ambil beberapa orang dengan lebih dari satu shift
                if len(available_names) < 3:
                    messagebox.showerror("ERROR", f"Not enough people available for day {current_day} ({day + 1}). Please check availability and leaves.")
                    return

                # Urutkan orang berdasarkan jumlah total shift yang sudah dimiliki
                available_names.sort(key=lambda name: total_shift_count[name])

                # Pilih 3 orang untuk shift hari tersebut secara merata
                day_shift = available_names[:3]

                # Update shift_count untuk setiap shift
                shift_count[day_shift[0]]["Malam"] += 1
                shift_count[day_shift[1]]["Pagi"] += 1
                shift_count[day_shift[2]]["Sore"] += 1

                # Update total shift per orang
                total_shift_count[day_shift[0]] += 1
                total_shift_count[day_shift[1]] += 1
                total_shift_count[day_shift[2]] += 1

                # Simpan shift per minggu
                current_week = (day // 7) + 1
                weekly_shift_count[day_shift[0]]["Malam"].append(current_week)
                weekly_shift_count[day_shift[1]]["Pagi"].append(current_week)
                weekly_shift_count[day_shift[2]]["Sore"].append(current_week)

                # Tambahkan nama orang yang cuti ke shift data dengan keterangan 'On Leave'
                for name in leave_names:
                    day_shift.append(f"{name} on leave")

                # Simpan data shift untuk hari tersebut
                shift_data.append([current_day, day + 1] + day_shift)

            # Membuat workbook dan worksheet baru
            global workbook, sheet
            workbook = Workbook()  # Workbook baru
            sheet = workbook.active
            sheet.title = "Shift Data"

            # Menulis header
            sheet.append(["Hari", "Tanggal", "Malam", "Pagi", "Sore"])
            cell_bold(1, 5)

            # Warna merah muda untuk hari Sabtu dan Minggu
            pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")

            # Menulis data shift harian
            for i, day_shift in enumerate(shift_data):
                day_name = day_shift[0]
                sheet.append(day_shift)

                # Dapatkan baris terakhir yang baru saja ditambahkan
                last_row = sheet.max_row

                # Jika hari Sabtu atau Minggu, warnai baris tersebut dengan merah muda
                if day_name == 'Saturday' or day_name == 'Sunday':
                    for cell in sheet[last_row][:5]:
                        cell.fill = pink_fill

            # Menambahkan dua baris kosong sebagai pemisah
            sheet.append([])
            sheet.append([])
            sheet.append([""])

            # Tambahkan teks
            sheet.append(["Total Shift per-weeks"])
            merged_center_bold()

            # Menulis total shift per minggu
            current_week = 1
            while current_week <= (max_days // 7) + 1:
                sheet.append([f"Week {current_week}"])
                merged_center_bold()
                sheet.append(["", "Night", "Morning", "Afternoon"])

                for name, shifts in weekly_shift_count.items():
                    night_shifts = shifts["Malam"].count(current_week)
                    morning_shifts = shifts["Pagi"].count(current_week)
                    afternoon_shifts = shifts["Sore"].count(current_week)
                    sheet.append([name, night_shifts, morning_shifts, afternoon_shifts])

                sheet.append([])  # Pemisah antar minggu
                current_week += 1

            # Menambahkan dua baris kosong
            sheet.append([])
            sheet.append(["Total Shift per-month"])  # Tambahkan teks
            merged_center_bold()
            sheet.append(["Shifts", "Night", "Morning", "Afternoon","Total"])

            # Menulis hasil shift untuk setiap orang dalam format tabel
            for name, shifts in shift_count.items():
                total_shifts = shifts["Malam"] + shifts["Pagi"] + shifts["Sore"]
                sheet.append([name, shifts["Malam"], shifts["Pagi"], shifts["Sore"], total_shifts])

            # Simpan file Excel
            filename = "shift.xlsx"
            workbook.save(filename)

            # Buka file Excel secara otomatis
            os.startfile(filename)

        else:
            messagebox.showerror("ERROR", "No names to save!")

# Fungsi tanpa aturan cuti
def withoutrulescuti():
    global max_days
    if hasattr(root, 'new_window') and root.new_window.winfo_exists():
        messagebox.showerror("WARNING", "Save or close settings first!", parent=root.new_window)

    else:
        max_days = int(root.datecomBox.get())
        start_day = root.daycomBox.get()

        days_of_week = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        start_index = days_of_week.index(start_day)
        week_days = days_of_week[start_index:] + days_of_week[:start_index]

        if inputted_names:
            # Shuffle sekali sebelum loop untuk efisiensi
            random.shuffle(inputted_names)

            shift_data = []
            total_names = len(inputted_names)

            # Dictionary untuk menghitung jumlah shift
            shift_count = {name: {"Malam": 0, "Pagi": 0, "Sore": 0} for name in inputted_names}

            # Dictionary untuk menghitung shift per minggu
            weekly_shift_count = {name: {"Malam": [], "Pagi": [], "Sore": []} for name in inputted_names}

            for day in range(max_days):
                current_day = week_days[day % len(week_days)]

                # Ambil nama secara berurutan tanpa pengacakan tambahan
                day_shift = [inputted_names[i % total_names] for i in range(day * 3, (day + 1) * 3)]

                # Update shift_count untuk setiap shift
                shift_count[day_shift[0]]["Malam"] += 1
                shift_count[day_shift[1]]["Pagi"] += 1
                shift_count[day_shift[2]]["Sore"] += 1

                # Simpan shift per minggu
                current_week = (day // 7) + 1  # Menghitung minggu keberapa
                weekly_shift_count[day_shift[0]]["Malam"].append(current_week)
                weekly_shift_count[day_shift[1]]["Pagi"].append(current_week)
                weekly_shift_count[day_shift[2]]["Sore"].append(current_week)

                shift_data.append([current_day, day + 1] + day_shift)

            # Membuat workbook dan worksheet baru
            global workbook, sheet
            workbook = Workbook()  # Workbook baru
            sheet = workbook.active
            sheet.title = "Shift Data"

            # Menulis header
            sheet.append(["Hari", "Tanggal", "Malam", "Pagi", "Sore"])
            cell_bold(1, 5)

            # Warna merah muda untuk hari Sabtu dan Minggu
            pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")

            # Menulis data shift harian
            for i, day_shift in enumerate(shift_data):
                day_name = day_shift[0]
                sheet.append(day_shift)

                # Dapatkan baris terakhir yang baru saja ditambahkan
                last_row = sheet.max_row

                # Jika hari Sabtu atau Minggu, warnai baris tersebut dengan merah muda
                if day_name == 'Saturday' or day_name == 'Sunday':
                    for cell in sheet[last_row]:
                        cell.fill = pink_fill

            # Menambahkan dua baris kosong sebagai pemisah
            sheet.append([])
            sheet.append([])
            sheet.append([""])

            # Tambahkan teks
            sheet.append(["Total Shift per-weeks"])
            merged_center_bold()

            # Menulis total shift per minggu
            current_week = 1
            while current_week <= (max_days // 7) + 1:
                sheet.append([f"Week {current_week}"])
                merged_center_bold()
                sheet.append(["", "Night", "Morning", "Afternoon"])

                for name, shifts in weekly_shift_count.items():
                    night_shifts = shifts["Malam"].count(current_week)
                    morning_shifts = shifts["Pagi"].count(current_week)
                    afternoon_shifts = shifts["Sore"].count(current_week)
                    sheet.append([name, night_shifts, morning_shifts, afternoon_shifts])

                sheet.append([])  # Pemisah antar minggu
                current_week += 1

            # Menambahkan dua baris kosong
            sheet.append([])
            sheet.append(["Total Shift per-month"])  # Tambahkan teks
            merged_center_bold()
            sheet.append(["Shifts", "Night", "Morning", "Afternoon"])

            # Menulis hasil shift untuk setiap orang dalam format tabel
            for name, shifts in shift_count.items():
                sheet.append([name, shifts["Malam"], shifts["Pagi"], shifts["Sore"]])

            # Simpan file Excel
            filename = "shift.xlsx"
            workbook.save(filename)

            # Buka file Excel secara otomatis
            os.startfile(filename)

        else:
            messagebox.showerror("ERROR", "No names to save!")

def update_font(event):
    # Menghitung ukuran font berdasarkan lebar jendela
    font_size = int(root.winfo_width() / 32) 

    # Update font untuk semua widget yang sesuai
    root.logFrame.config(font=("Helvetica", font_size))
    root.logFrame.config(font=("Helvetica", font_size))

# Fungsi merged cell dan bold alignment tengah
def merged_center_bold():
    bold_font = Font(bold=True)
    # Dapatkan baris terakhir
    last_row = sheet.max_row

    # Merge cell dari kolom 1 sampai kolom 4 (misal, menyesuaikan dengan kebutuhan)
    sheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=4)

    # Terapkan gaya bold dan center alignment pada cell
    merged_cell = sheet.cell(row=last_row, column=1)
    merged_cell.font = bold_font
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")

def cell_bold(start_col, end_col):
    last_row = sheet.max_row

    for col in range(start_col, end_col + 1):  # Gunakan +1 karena range() tidak menyertakan end_col
        cell = sheet.cell(row=last_row, column=col)
        cell.font = bold_font

def validate_day_input(event, entry_widget):
    # Ambil teks yang dimasukkan di entry
    max_days1 = int(root.datecomBox.get())  # Ambil nilai max days dari combobox
    user_input = entry_widget.get()

    # Mengizinkan backspace, tab, dan enter untuk tidak memvalidasi
    if event.keysym in ['BackSpace', 'Tab', 'Return']:
        return

    # Memeriksa jika input valid: angka 1-2 digit yang dipisahkan oleh semicolon, termasuk input sementara seperti '12;'
    if re.match(r"^([0-9]{1,2})(;[0-9]{1,2})*;?$", user_input):  # Memeriksa pola angka-angka yang dipisahkan oleh ;, termasuk yang diakhiri ;
        # Pisahkan input berdasarkan semicolon, abaikan jika kosong terakhir
        days = [day for day in user_input.split(';') if day]

        # Validasi setiap angka dalam list days
        for day in days:
            if not (1 <= int(day) <= max_days1):  # Pastikan setiap angka berada dalam rentang yang valid
                entry_widget.delete(0, "end")
                messagebox.showerror("WARNING", f"Invalid date {day}! Please input a valid date between 1 and {max_days1}.", parent=root.new_window)
                return
        # Input valid, biarkan tetap ada
        pass
    else:
        # Hapus input jika format tidak valid
        entry_widget.delete(0, "end")
        messagebox.showerror("WARNING", f"Input must be numbers between 1 and {max_days1}, separated by semicolons (;).", parent=root.new_window)

def cancel_button():
    root.new_window.destroy()

def save_button():
    # Dictionary untuk menyimpan data yang di-input
    global saved_data
    saved_data = {}

    # Loop melalui setiap nama yang di-input
    for index, name in enumerate(inputted_names):
        # Ambil nilai dari entry untuk total days dalam format 14;15 atau 13
        total_leaves_input = root.totalleaves[name].get()  # Mengambil nilai dari CTkEntry yang sesuai dengan nama
        # Pisahkan input berdasarkan pemisah ';' untuk menghasilkan list day_entries
        day_entries = total_leaves_input.split(';')  # Split input menjadi list berdasarkan ';'
        if total_leaves_input == '':
            total_days = 0
        else:
            total_days = len(day_entries)
        # Simpan data ke dalam dictionary dengan nama sebagai kunci
        saved_data[name] = {
            "total_leaves": total_days,  # Menyimpan total hari dalam format string
            "days": day_entries  # Simpan list day_entries yang dipecah
        }

    # Cetak data yang tersimpan untuk memastikan
    print("Data yang disimpan:", saved_data)

    # Setelah menyimpan, Anda bisa menambahkan logika untuk menyimpan data ke database, file, dsb.
    messagebox.showinfo("Success", "Data successfully saved!", parent=root.new_window)
    root.new_window.destroy()

# Window popup rules
def top_level_win():
    text_content = root.logFrame.get("1.0", END).strip()
    if not text_content:
        messagebox.showerror("WARNING", "Input names first!", parent=root)
    else: 
        # Cek jika jendela baru sudah ada, jika belum, buat yang baru
        if not hasattr(root, 'new_window') or not root.new_window.winfo_exists():
            root.new_window = CTkToplevel(root)
            root.new_window.title("Shift Configuration")
            root.new_window.resizable(False, False)
            root.new_window.geometry("400x305")
            root.new_window.attributes("-topmost", True)

            # frame scrollable
            root.infolabel = CTkLabel(root.new_window, text="*Input dates for annual leaves \neg. 14, 14;15, 14;15;16. For multiple day, use (;) as separator!")
            root.infolabel.pack(pady=5)
            root.my_frame = CTkScrollableFrame(root.new_window, width=400, height=200)
            root.my_frame.pack(pady=5)

            root.totalleaves = {}
            
            # widget frame untuk setiap nama
            for index, name in enumerate(inputted_names):
                # Display the name in a row
                root.namelabel = CTkLabel(root.my_frame, text=name, bg_color='#0C4E8A', text_color='white', width=400, font=("Arial", 12, "bold"))
                root.namelabel.grid(row=index*6, columnspan=30, pady=5, sticky="w")

                # Label untuk annual leaves di bawah nama
                root.annuallabel = CTkLabel(root.my_frame, text="Annual leaves:")
                root.annuallabel.grid(row=index*6+1, column=0, pady=5, sticky="w")

                # Spinbox atau CTkEntry untuk memasukkan total annual leaves
                root.totalleaves[name] = CTkEntry(root.my_frame, width=100)  # Simpan widget ke dictionary
                root.totalleaves[name].grid(row=index*6+1, column=1, pady=5, sticky="w")
                
                # Bind untuk memvalidasi input agar hanya angka
                root.totalleaves[name].bind("<KeyRelease>", lambda event, entry_widget=root.totalleaves[name]: validate_day_input(event, entry_widget))

            # save Button
            root.saveBTN = CTkButton(root.new_window, text="Save", command=save_button, width=195, fg_color="green", corner_radius=32, text_color='white')
            root.saveBTN.pack(side = LEFT, expand=TRUE,pady=5)

            # cancel Button
            root.cancelBTN = CTkButton(root.new_window, text="Cancel", command=cancel_button, width=195, fg_color="red", corner_radius=32, text_color='white')
            root.cancelBTN.pack(side = LEFT, expand=TRUE,pady=5)

            # Fokus ke jendela baru
            root.new_window.focus()
        else:
            # Jika jendela sudah ada, fokus ke jendela tersebut
            root.new_window.focus()

# Buat objek class tk
root = CTk()

# Setting ukuran judul
root.title("Managed Service Shift Generator")
root.geometry("360x350")
root.resizable(False, False)
set_appearance_mode("light")
root.bind("<Configure>", update_font)

# Variable Sheets
bold_font = Font(bold=True)

# Variable
saved_data = {}
namevar = StringVar()
inputted_names = []

create_widgets()
root.mainloop()